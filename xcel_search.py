#!/usr/bin/env python3
"""
Search all .txt, .docx, and .xlsx files under a directory tree for words or phrases.

Run the script, then:
  1. Enter the root directory to search.
  2. Enter a **comma-separated** list of words/phrases (case-insensitive).

Every result shown contains **all** of the supplied terms.
"""

import os
from typing import List
from docx import Document
from openpyxl import load_workbook


def file_contains_terms(path: str, terms: List[str]) -> bool:
    """Return True iff *all* terms occur (case-insensitive) in the file text."""
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".txt":
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read().lower()

        elif ext == ".docx":
            doc = Document(path)
            text = "\n".join(p.text for p in doc.paragraphs).lower()

        elif ext == ".xlsx":
            wb = load_workbook(filename=path, data_only=True)
            text_parts: List[str] = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if isinstance(cell, str):
                            text_parts.append(cell.lower())
            text = "\n".join(text_parts)

        else:
            return False  # Unsupported extension

    except Exception as exc:
        print(f"Warning: failed to read {path!r}: {exc}")
        return False

    return all(term in text for term in terms)


def prompt_directory() -> str:
    """Ask the user for a directory path until a valid one is entered."""
    while True:
        directory = input("Enter the root directory to search: ").strip().strip('"')
        if os.path.isdir(directory):
            return directory
        print("That directory does not exist. Please try again.\n")


def prompt_terms() -> List[str]:
    """Prompt for a comma-separated list of words/phrases and return the list."""
    while True:
        raw = input("Enter words/phrases to search for (comma-separated): ").strip()
        terms = [part.strip().lower() for part in raw.split(',') if part.strip()]
        if terms:
            return terms
        print("You must enter at least one term.\n")


def main():
    directory = prompt_directory()
    terms = prompt_terms()

    matches: List[str] = []
    for root, _, files in os.walk(directory):
        for fname in files:
            if os.path.splitext(fname)[1].lower() in (".txt", ".docx", ".xlsx"):
                fullpath = os.path.join(root, fname)
                if file_contains_terms(fullpath, terms):
                    matches.append(fullpath)

    if matches:
        print("\nFiles containing all terms:")
        for m in matches:
            print(m)
    else:
        print("No files found containing all of those terms.")


if __name__ == "__main__":
    main()
