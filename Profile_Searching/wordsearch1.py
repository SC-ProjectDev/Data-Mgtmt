#!/usr/bin/env python3
"""
Search .txt, .docx, .xlsx, .xls, and .pdf files under a directory tree for words or phrases.

Run the script, then:
  1. Enter the root directory to search.
  2. Enter a comma-separated list of words/phrases (case-insensitive).

Every result shown contains all of the supplied terms.
"""

import os
from typing import List
from docx import Document
from openpyxl import load_workbook
from PyPDF2 import PdfReader
import xlrd  # For .xls files


def file_contains_terms(path: str, terms: List[str]) -> bool:
    """Return True iff all terms occur (case-insensitive) in the file text."""
    ext = os.path.splitext(path)[1].lower()
    text = ""

    try:
        if ext == ".txt":
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()

        elif ext == ".docx":
            doc = Document(path)
            text = "\n".join(p.text for p in doc.paragraphs)

        elif ext == ".xlsx":
            wb = load_workbook(path, read_only=True, data_only=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            text += f"{cell} "

        elif ext == ".xls":
            book = xlrd.open_workbook(path)
            for sheet in book.sheets():
                for row_idx in range(sheet.nrows):
                    for col_idx in range(sheet.ncols):
                        val = sheet.cell_value(row_idx, col_idx)
                        text += f"{val} "

        elif ext == ".pdf":
            reader = PdfReader(path)
            for page in reader.pages:
                text += page.extract_text() or ""

        else:
            return False  # Unsupported extension

    except Exception as exc:
        print(f"Warning: failed to read {path!r}: {exc}")
        return False

    text = text.lower()
    return all(term in text for term in terms)


def prompt_directory() -> str:
    while True:
        directory = input("Enter the root directory to search: ").strip().strip('"')
        if os.path.isdir(directory):
            return directory
        print("That directory does not exist. Please try again.\n")


def prompt_terms() -> List[str]:
    while True:
        raw = input("Enter words/phrases to search for (comma-separated): ").strip()
        terms = [part.strip().lower() for part in raw.split(',') if part.strip()]
        if terms:
            return terms
        print("You must enter at least one term.\n")


def main():
    directory = prompt_directory()
    terms = prompt_terms()

    matches: List[str] = []
    for root, _, files in os.walk(directory):
        for fname in files:
            if os.path.splitext(fname)[1].lower() in (".txt", ".docx", ".xlsx", ".xls", ".pdf"):
                fullpath = os.path.join(root, fname)
                if file_contains_terms(fullpath, terms):
                    matches.append(fullpath)

    if matches:
        print("\nFiles containing all terms:")
        for m in matches:
            print(m)
    else:
        print("No files found containing all of those terms.")


if __name__ == "__main__":
    main()
