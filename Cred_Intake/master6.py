#!/usr/bin/env python3
"""
Script to extract and categorize tables from multiple Excel workbooks into a single master file,
creating separate category sheets and preserving IPA column styles,
and logging processing details.

Usage:
    python excel_master_single_output.py /path/to/input.xlsx_or_dir /path/to/output_dir
"""

import argparse
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from copy import copy
import getpass
from datetime import datetime

# Define the lists of IPA organization columns (styled columns)
IPA_COLUMNS = [
    'Access Primary Care Medical Group',
    'Accountable Health Care',
    'Allied Pacific of California',
    'All American Medical Group',
    'Alpha Care Medical Group',
    'American Acupuncture Chinese Medicine',
    'Associated Hispanic Physicians',
    'Arroyo Vista Family Health Clinic',
    'Bay Area Care Partners',
    'Beverly Alianza IPA',
    'Central Valley Medical Group',
    'Community Family Care IPA',
    'Diamond Bar Medical Group',
    'For Your Benefit',
    'Hana Hou Medical Group',
    'Central California Physician Partners',
    'Jade Health IPA',
]

# Define common columns and their aliases
COMMON_COLUMNS = [
    ['COMMITTEE (Astrana Health)'],
    ['LEVEL       (1 or 2)'],
    ['LAST/HDO Name'],
    ['FIRST'],
    ['DEGREE'],
    ['TYPE (PCP/SCP/AHP/HDO)'],
    ['SPECIALTY'],
    ['NPI'],
    ['LICENSE'],
    ['BOARD CERTIFICATION'],
    ['BOARD CERTIFICATION EXP.'],
    ['Vendor'],
    ['COUNTY'],
    ['NEW CONTRACT(Y/N) OR PREVIOUS CRED DATE FOR RECREDS'],
    ['COMMITTEE SUBMISSION DATE'],
    ['STATUS'],
    ['Credentialing Coordinator/ Specialist'],
    ['CAQH'],
    ['GEMINI'],
]

# Define the target headers per category
target_headers = {
    'Initial':        [[col] for col in IPA_COLUMNS] + COMMON_COLUMNS,
    'Recreds':        [[col] for col in IPA_COLUMNS] + COMMON_COLUMNS,
    'HDOs':           [[col] for col in IPA_COLUMNS] + COMMON_COLUMNS,
    'Links':          [[col] for col in IPA_COLUMNS] + COMMON_COLUMNS,
    'Reinstatements': [[col] for col in IPA_COLUMNS] + COMMON_COLUMNS,
}

# Build a mapping from lowercase alias to canonical header
alias_map = {}
for cat, alias_lists in target_headers.items():
    for aliases in alias_lists:
        canonical = aliases[0]
        for alias in aliases:
            alias_map[alias.strip().lower()] = canonical


def detect_category(status_val):
    """
    Determine category based on STATUS value.
    """
    if pd.notna(status_val):
        val = str(status_val).strip().lower()
        if 'reinstatemen' in val:
            return 'Reinstatements'
        if 'initial' in val:
            return 'Initial'
        if 'recred' in val:
            return 'Recreds'
        if 'hdo' in val:
            return 'HDOs'
        if 'link' in val:
            return 'Links'
    return None


def is_header_row(row):
    texts = [str(cell.value).strip().lower() for cell in row if cell.value]
    return 'status' in texts and any(col.lower() in texts for col in IPA_COLUMNS)


def parse_master(input_file: str):
    """
    Parse a single workbook and collect data rows by category.
    Returns a dict: {category: [(row_data, style_map), ...], ...}
    """
    wb = load_workbook(input_file, data_only=True)
    ws = wb.active

    collected = {cat: [] for cat in target_headers}
    rows = list(ws.iter_rows())
    i = 0

    while i < len(rows):
        row = rows[i]
        # Skip empty rows
        if all(cell.value is None for cell in row):
            i += 1
            continue
        # Detect header
        if is_header_row(row):
            header_map = {}
            for idx, cell in enumerate(row, start=1):
                if cell.value and isinstance(cell.value, str):
                    key = cell.value.strip().lower()
                    if key in alias_map:
                        header_map[alias_map[key]] = idx
            j = i + 1
            while j < len(rows) and any(c.value is not None for c in rows[j]):
                status_cell = rows[j][header_map.get('STATUS')-1] if 'STATUS' in header_map else None
                status_val = status_cell.value if status_cell else None
                category = detect_category(status_val)
                if category:
                    row_data = {}
                    style_map = {}
                    for canonical, col_idx in header_map.items():
                        cell = ws.cell(row=j+1, column=col_idx)
                        row_data[canonical] = cell.value
                        if canonical in IPA_COLUMNS:
                            style_map[canonical] = (cell.fill, cell.font, cell.border)
                    collected[category].append((row_data, style_map))
                j += 1
            i = j
        else:
            i += 1
    return collected


def write_master(output_file: str, collected: dict):
    """
    Write the combined collected data into one master Excel file,
    one sheet per category, preserving IPA column styles.
    """
    # Write dataframes
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for cat, rows in collected.items():
            if not rows:
                continue
            data_dicts = [r for r, _ in rows]
            df = pd.DataFrame(data_dicts)
            desired = [aliases[0] for aliases in target_headers[cat]]
            cols = [c for c in desired if c in df.columns]
            df.reindex(columns=cols).to_excel(writer, sheet_name=cat, index=False)

    # Reopen and apply styles
    wb2 = load_workbook(output_file)
    for cat, rows in collected.items():
        if not rows or cat not in wb2.sheetnames:
            continue
        ws2 = wb2[cat]
        header_row = next(ws2.iter_rows(min_row=1, max_row=1))
        dest_header = {cell.value: cell.column for cell in header_row}
        for idx, (_r, style_map) in enumerate(rows, start=2):
            for col, (fill, font, border) in style_map.items():
                if col in dest_header:
                    dest_cell = ws2.cell(row=idx, column=dest_header[col])
                    dest_cell.fill = copy(fill)
                    dest_cell.font = copy(font)
                    dest_cell.border = copy(border)
    wb2.save(output_file)


def log_to_excel(log_file: str, log_data: list):
    """Append processing logs to an Excel log sheet."""
    log_path = Path(log_file)
    if log_path.exists():
        book = load_workbook(log_file)
        writer = pd.ExcelWriter(log_file, engine='openpyxl', mode='a', if_sheet_exists='overlay')
    else:
        writer = pd.ExcelWriter(log_file, engine='openpyxl')
        book = writer.book
    df_log = pd.DataFrame(log_data)
    startrow = book['Log'].max_row if 'Log' in book.sheetnames else 0
    df_log.to_excel(
        writer,
        sheet_name='Log',
        index=False,
        header=startrow == 0,
        startrow=startrow
    )
    writer.close()


def main():
    parser = argparse.ArgumentParser(
        description='Aggregate multiple Excel files into one master categorized file with logging.'
    )
    parser.add_argument('input_path', help='Path to source Excel file or directory containing Excel files')
    parser.add_argument('output_dir', help='Directory to save output and log files')
    args = parser.parse_args()

    inp = Path(args.input_path)
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    user = getpass.getuser()
    timestamp = lambda: datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_file = out_dir / "processing_log.xlsx"
    log_records = []

    # Prepare combined collector
    combined = {cat: [] for cat in target_headers}

    # Gather files
    if inp.is_file():
        files = [inp]
    elif inp.is_dir():
        files = [f for f in inp.glob('*.xlsx') if not f.name.startswith('~$')]
    else:
        print(f"Invalid input: {inp}")
        return

    for file in files:
        print(f"Processing {file.name}...")
        try:
            data = parse_master(str(file))
            for cat in combined:
                combined[cat].extend(data[cat])
            log_records.append({
                'Timestamp': timestamp(),
                'Username': user,
                'Input File': file.name,
                'Sheets Extracted': ', '.join([cat for cat, rows in data.items() if rows]),
                'Success': True,
                'Message': 'Merged'
            })
        except Exception as e:
            log_records.append({
                'Timestamp': timestamp(),
                'Username': user,
                'Input File': file.name,
                'Sheets Extracted': '',
                'Success': False,
                'Message': str(e)
            })

    # Write single master file
    master_file = out_dir / "ALL_MASTER.xlsx"
    write_master(str(master_file), combined)
    print(f"Master file written to '{master_file}'")

    # Log results
    if log_records:
        log_to_excel(str(log_file), log_records)
        print(f"Log written to '{log_file}'")

if __name__ == '__main__':
    main()
