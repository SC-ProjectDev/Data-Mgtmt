import os
import re
import pdfplumber
import openpyxl
from openpyxl import Workbook

# Define the full list of IPA headers
IPA_HEADERS = [
    "NPI",
    "Access Primary Care Medical Group",
    "Accountable Health Care",
    "Allied Pacific of California",
    "All American Medical Group",
    "Alpha Care Medical Group",
    "American Acupuncture Chinese Medicine",
    "Associated Hispanic Physicians",
    "Arroyo Vista Family Health Clinic",
    "Bay Area Care Partners",
    "Beverly Alianza IPA",
    "Central Valley Medical Group",
    "Community Family Care IPA",
    "Diamond Bar Medical Group",
    "For Your Benefit",
    "Hana Hou Medical Group",
    "Central California Physician Partners",
    "Jade Health IPA",
    "Filenames Found"
]

def extract_npi_and_networks(pdf_path):
    npi = None
    networks = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                lines = page.extract_text().splitlines()
                for line in lines:
                    # Extract NPI
                    npi_match = re.search(r"NPI:\s*(\d{10})", line)
                    if npi_match:
                        npi = npi_match.group(1)
                    # Extract Network
                    if "IPA Medical Group(s):" in line:
                        idx = lines.index(line)
                        if idx + 1 < len(lines):
                            networks.append(lines[idx + 1].strip())
    except Exception as e:
        print(f"Failed to read {pdf_path}: {e}")
    return npi, networks

def collect_pdf_data(pdf_dir):
    pdf_data = {}
    for root, _, files in os.walk(pdf_dir):
        for file in files:
            if file.lower().endswith(".pdf"):
                path = os.path.join(root, file)
                npi, networks = extract_npi_and_networks(path)
                if npi:
                    if npi not in pdf_data:
                        pdf_data[npi] = {"networks": set(), "files": []}
                    pdf_data[npi]["networks"].update(networks)
                    pdf_data[npi]["files"].append(file)
    return pdf_data

def load_npis_from_excel(sheet_path, column):
    wb = openpyxl.load_workbook(sheet_path, data_only=True)
    ws = wb.active
    npis = []
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column_letter == column and cell.value:
                npis.append(str(cell.value).strip())
    return npis

def generate_output_excel(npis, pdf_data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.append(IPA_HEADERS)

    for npi in npis:
        row = [npi]
        found_networks = pdf_data.get(npi, {}).get("networks", set())
        for header in IPA_HEADERS[1:-1]:  # Skip NPI and Filenames Found
            row.append("X" if header in found_networks else "")
        filenames = ", ".join(pdf_data.get(npi, {}).get("files", []))
        row.append(filenames)
        ws.append(row)

    wb.save(output_path)

def main():
    pdf_dir = input("Enter the directory containing PDF files: ").strip()
    excel_path = input("Enter the path to the Excel file: ").strip()
    column = input("Enter the column letter for the NPIs (e.g., CL): ").strip().upper()

    print("\nScanning PDFs... This may take a moment...")
    pdf_data = collect_pdf_data(pdf_dir)

    print("\nReading NPIs from spreadsheet...")
    npis = load_npis_from_excel(excel_path, column)

    print("\nGenerating output spreadsheet...")
    output_path = os.path.splitext(excel_path)[0] + "_output_profiles.xlsx"
    generate_output_excel(npis, pdf_data, output_path)
    print(f"\nDone! Results saved to: {output_path}")

if __name__ == "__main__":
    main()
